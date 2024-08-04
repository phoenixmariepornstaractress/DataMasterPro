import argparse
import json
import logging
import shutil
import smtplib
from datetime import datetime
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
import sqlite3
from cryptography.fernet import Fernet
from nltk.sentiment.vader import SentimentIntensityAnalyzer
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from PIL import Image
from sklearn.cluster import KMeans
from sklearn.ensemble import IsolationForest
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error, r2_score
from sklearn.model_selection import GridSearchCV, train_test_split
from sklearn.preprocessing import OneHotEncoder, StandardScaler
from stable_baselines3 import PPO
from stable_baselines3.common.env_util import make_vec_env
from statsmodels.tsa.holtwinters import ExponentialSmoothing
from surprise import Dataset, Reader, SVD, accuracy
from surprise.model_selection import train_test_split as surprise_train_test_split
from tensorflow.keras import layers, models, preprocessing
from tensorflow.keras.models import Sequential
from tensorflow.keras.preprocessing import image
from transformers import pipeline
from tpot import TPOTRegressor
import shap
import spacy
import torchvision.models as models
import torchvision.transforms as transforms

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def backup_excel(file_path):
    backup_path = f"{file_path}.backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    shutil.copy(file_path, backup_path)
    logging.info(f"Backup created: {backup_path}")

def send_email_notification(config, subject, body, attachment_path=None):
    from_email = config['email']['from_email']
    from_password = config['email']['from_password']
    to_email = config['email']['to_email']

    msg = MIMEMultipart()
    msg['From'] = from_email
    msg['To'] = to_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))

    if attachment_path:
        with open(attachment_path, 'rb') as f:
            part = MIMEApplication(f.read(), Name=attachment_path)
            part['Content-Disposition'] = f'attachment; filename="{attachment_path}"'
            msg.attach(part)

    try:
        server = smtplib.SMTP(config['email']['smtp_server'], config['email']['smtp_port'])
        server.starttls()
        server.login(from_email, from_password)
        server.sendmail(from_email, to_email, msg.as_string())
        server.quit()
        logging.info(f"Email sent to {to_email}")
    except Exception as e:
        logging.error(f"Failed to send email: {e}")

def validate_data(df):
    if df.isnull().values.any():
        logging.warning("Data contains null values. Filling missing values with median.")
        df.fillna(df.median(), inplace=True)
    return df

def handle_outliers(df, threshold=3):
    z_scores = np.abs((df - df.mean()) / df.std())
    df = df[(z_scores < threshold).all(axis=1)]
    return df

def transform_data(df):
    if 'name' in df.columns:
        df['name'] = df['name'].str.upper()
    return df

def format_excel(file_path, sheet_name):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    wb.save(file_path)

def encrypt_data(df, encryption_key):
    fernet = Fernet(encryption_key)
    for column in df.columns:
        if df[column].dtype == 'object':  # Encrypt string columns
            df[column] = df[column].apply(lambda x: fernet.encrypt(x.encode()).decode() if x else x)
    return df

def compare_data(old_df, new_df):
    diff = new_df.compare(old_df)
    return diff

def feature_engineering(df, feature_columns):
    for column in feature_columns:
        if df[column].dtype in [np.int64, np.float64]:
            df[f'{column}_log'] = np.log1p(df[column])
            df[f'{column}_sqrt'] = np.sqrt(df[column])
    return df

def evaluate_model(model, X_test, y_test):
    y_pred = model.predict(X_test)
    mse = mean_squared_error(y_test, y_pred)
    r2 = r2_score(y_test, y_pred)
    logging.info(f'Model Evaluation:\nMSE: {mse}\nR^2: {r2}')
    return mse, r2

def plot_feature_importance(model, feature_columns):
    plt.figure(figsize=(10, 6))
    feature_importance = pd.Series(model.coef_, index=feature_columns)
    feature_importance.nlargest(10).plot(kind='barh')
    plt.title('Top 10 Feature Importances')
    plt.show()

def plot_predictions(y_test, y_pred):
    plt.figure(figsize=(10, 6))
    plt.scatter(y_test, y_pred, alpha=0.3)
    plt.xlabel('Actual Values')
    plt.ylabel('Predicted Values')
    plt.title('Actual vs Predicted Values')
    plt.show()

def hyperparameter_tuning(X_train, y_train):
    param_grid = {'fit_intercept': [True, False], 'normalize': [True, False]}
    grid_search = GridSearchCV(LinearRegression(), param_grid, cv=5, scoring='neg_mean_squared_error')
    grid_search.fit(X_train, y_train)
    best_params = grid_search.best_params_
    logging.info(f'Best Hyperparameters: {best_params}')
    return best_params

def automl_train(X_train, y_train):
    tpot_model = TPOTRegressor(generations=5, population_size=20, verbosity=2)
    tpot_model.fit(X_train, y_train)
    return tpot_model

def train_deep_learning_model(X_train, y_train, X_test, y_test):
    model = Sequential()
    model.add(layers.Dense(64, input_dim=X_train.shape[1], activation='relu'))
    model.add(layers.Dropout(0.5))
    model.add(layers.Dense(32, activation='relu'))
    model.add(layers.Dense(1, activation='linear'))

    model.compile(optimizer='adam', loss='mse', metrics=['mae'])
    model.fit(X_train, y_train, epochs=50, batch_size=32, validation_data=(X_test, y_test))

    return model

def explain_model_predictions(model, X_train):
    explainer = shap.Explainer(model)
    shap_values = explainer(X_train)
    shap.summary_plot(shap_values, X_train)

def train_predictive_model(df, target_column, feature_columns, tune_hyperparameters, use_automl):
    X = df[feature_columns]
    y = df[target_column]

    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    if use_automl:
        tpot_model = automl_train(X_train, y_train)
        y_pred = tpot_model.predict(X_test)
        mse = mean_squared_error(y_test, y_pred)
        r2 = r2_score(y_test, y_pred)
        logging.info(f'AutoML Model Evaluation:\nMSE: {mse}\nR^2: {r2}')
        tpot_model.export('tpot_best_model.py')
    else:
        if tune_hyperparameters:
            best_params = hyperparameter_tuning(X_train, y_train)
            model = LinearRegression(**best_params)
        else:
            model = LinearRegression()

        model.fit(X_train, y_train)
        mse, r2 = evaluate_model(model, X_test, y_test)
        plot_feature_importance(model, feature_columns)
        plot_predictions(y_test, model.predict(X_test))

    deep_model = train_deep_learning_model(X_train, y_train, X_test, y_test)
    explain_model_predictions(model, X_train)

    return df, model

def detect_anomalies(df, columns, threshold=3):
    isolation_forest = IsolationForest(contamination=0.1)
    df['anomaly'] = isolation_forest.fit_predict(df[columns])
    anomalies = df[df['anomaly'] == -1]
    return anomalies

def autoencoder_anomaly_detection(df, columns):
    scaler = StandardScaler()
    scaled_df = scaler.fit_transform(df[columns])

    input_dim = scaled_df.shape[1]
    encoding_dim = 14

    input_layer = layers.Input(shape=(input_dim,))
    encoder = layers.Dense(encoding_dim, activation="relu")(input_layer)
    decoder = layers.Dense(input_dim, activation="sigmoid")(encoder)
    autoencoder = models.Model(inputs=input_layer, outputs=decoder)

    autoencoder.compile(optimizer='adam', loss='mse')
    autoencoder.fit(scaled_df, scaled_df, epochs=50, batch_size=32, shuffle=True, validation_split=0.2)

    predictions = autoencoder.predict(scaled_df)
    mse = np.mean(np.power(scaled_df - predictions, 2), axis=1)
    df['anomaly'] = mse > threshold

    anomalies = df[df['anomaly']]
    return anomalies

def parse_arguments():
    parser = argparse.ArgumentParser(description='Process some data.')
    parser.add_argument('--input', type=str, help='Path to the input Excel file')
    parser.add_argument('--output', type=str, help='Path to the output Excel file')
    parser.add_argument('--config', type=str, help='Path to the config file')
    parser.add_argument('--backup', action='store_true', help='Backup the Excel file before processing')
    parser.add_argument('--send_email', action='store_true', help='Send an email notification after processing')
    return parser.parse_args()

def main():
    args = parse_arguments()

    if args.config:
        with open(args.config, 'r') as config_file:
            config = json.load(config_file)

    df = pd.read_excel(args.input)
    df = validate_data(df)
    df = handle_outliers(df)
    df = transform_data(df)
    df = encrypt_data(df, config['encryption_key'])

    feature_columns = [col for col in df.columns if col != 'target_column']
    df, model = train_predictive_model(df, 'target_column', feature_columns, True, False)

    anomalies = detect_anomalies(df, feature_columns)
    logging.info(f"Detected anomalies:\n{anomalies}")

    autoencoder_anomalies = autoencoder_anomaly_detection(df, feature_columns)
    logging.info(f"Detected anomalies using autoencoder:\n{autoencoder_anomalies}")

    format_excel(args.output, 'Sheet1')

    df.to_excel(args.output, index=False)

    if args.send_email:
        send_email_notification(config, 'Data Processing Completed', 'The data processing task has been completed.', args.output)

    logging.info("Data processing completed successfully.")

if __name__ == '__main__':
    main()
